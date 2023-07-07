import tkinter

from tkinter import *

from tkinter.ttk import *

from tkinter import messagebox

import openpyxl

from openpyxl import Workbook

import pathlib

import sys

root=Tk()

root.geometry("800x800")

root.title("Registration Form")


file=pathlib.Path("backhand.xlsx")

if file.exists ():

    pass

else:

    file=Workbook()

    sheet=file.active

    sheet["A1"]="FULL NAME"

    sheet["B1"]="GENDER"

    sheet["C1"]="SUBJECT 1"

    sheet["D1"]="SUBJECT 2"

    sheet["E1"]="BIRTHDATE"

    file.save("backhand.xlsx")


#Label
label_0=Label(root,text="Registration Form",width=20,font=("bold",18))
label_0.place(x=90,y=60)

#Label+Entry
label_1=Label(root,text="Full Name:",width=10,font=("Times New Roman",12))
label_1.place(x=90,y=120)
entry_1=Entry(root)
entry_1.place(x=180,y=120)

#Radiobutton
label_2=Label(root,text="Gender:",width=10,font=("Times New Roman",12))
label_2.place(x=90,y=150)
var=IntVar()
Radiobutton(root,text="Male",width = 10, variable= var, value=1).place(x=180,y=150)
Radiobutton(root,text="Female",width = 10, variable= var, value=2).place(x=280,y=150)
Radiobutton(root,text="Other",width = 10, variable= var, value=3).place(x=380,y=150)

#Checkbutton
label_instr=Label(root,text="Choose any one from each row according to your subject choice:",width=100,font=("Times New Roman",12))
label_instr.place(x=90,y=180)
var1=IntVar()
Checkbutton(root,text=("Maths"),variable=var1).place(x=90,y=210)

var2=IntVar()                                                 
Checkbutton(root,text=("Biology"),variable=var2).place(x=160,y=210)

var3=IntVar()
Checkbutton(root,text=("CS"),variable=var3).place(x=90,y=230)

var4=IntVar()                                                 
Checkbutton(root,text=("PE"),variable=var4).place(x=160,y=230)

label_ask=Label(root,text="Enter Date of Birth:",width=90,font=("Times New Roman",12))
label_ask.place(x=90,y=250)

#SpinBox
my_spin_1=Spinbox(root, from_=1, to=31, width=10)
my_spin_1.place(x=90,y=270)

#Combobox
options = [
    "January",
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]
clicked=StringVar()
clicked.set(options[0])

drop=OptionMenu(root, clicked, *options)
drop.place(x=180,y=270)

my_spin_2=Spinbox(root, from_=1950, to=2021, width=10)
my_spin_2.place(x=270,y=270)

#Submit Button
def popup():
    messagebox.askyesno("Confirm Submission","You can only submit the form once. Confirm Submission?")
    a=entry_1.get()
    b=""
    tokenb=var.get()
    if tokenb==1:
        b="male"
    elif tokenb==2:
        b="female"
    else:
        b="other"

    c1=var1.get()
    if c1==1:
        c="Maths"
    else:
        c="Biology"

    d1=var3.get()
    if d1==1:
        d="CS"
    else:
        d="PE"

    e1=my_spin_1.get()
    e2=clicked.get()
    e3=my_spin_2.get()
    e=e1+" "+e2+" "+e3

    
    print("PLEASE WAIT WE ARE ADDING YOUR DETAILS IN THE SERVER...")

    file=openpyxl.load_workbook("backhand.xlsx")

    sheet=file.active

    sheet.cell(column=1, row=sheet.max_row+1, value=a)

    sheet.cell(column=2, row=sheet.max_row, value=b)

    sheet.cell(column=3, row=sheet.max_row, value=c)

    sheet.cell(column=4, row=sheet.max_row, value=d)

    sheet.cell(column=5, row=sheet.max_row, value=e)

    file.save("backhand.xlsx")

    print("YOUR DATA IS SUCCESSFULLY UPDATED THANK YOU!!")

    root.destroy()

bsub=Button(root, text="Submit", width=10, command=popup)
bsub.place(x=200,y=300)

root.mainloop
